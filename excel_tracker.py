# ============================================================
# SUIVI EXCEL DES RÉSERVATIONS — Restaurant Le Talier
# ============================================================

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

EXCEL_FILE = "reservations_le_talier.xlsx"

HEADERS = [
    "N° Résa", "Date résa", "Heure", "Nom client", "Téléphone",
    "Nb personnes", "Type service", "Commande", "Total (FCFA)",
    "Lien Google Calendar", "Date enregistrement"
]

HEADER_COLOR   = "1B5E20"   # Vert foncé
HEADER_FONT    = "FFFFFF"   # Blanc
ROW_ALT_COLOR  = "E8F5E9"   # Vert très clair (lignes paires)
ROW_BASE_COLOR = "FFFFFF"   # Blanc (lignes impaires)
TOTAL_COLOR    = "F1F8E9"   # Vert pâle


def _thin_border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _create_workbook():
    """Crée un nouveau classeur Excel avec l'en-tête et les styles."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Réservations"

    # ── Titre principal ──
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "🍽️  RESTAURANT LE TALIER — Suivi des Réservations & Commandes"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=HEADER_FONT)
    title_cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── En-têtes colonnes ──
    ws.append(HEADERS)
    header_row = ws[2]
    for cell in header_row:
        cell.font = Font(name="Arial", size=10, bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 24

    # ── Largeurs colonnes ──
    col_widths = [10, 14, 10, 22, 18, 14, 18, 50, 16, 40, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Feuille résumé ──
    ws_stat = wb.create_sheet("Statistiques")
    ws_stat["A1"] = "Statistiques Le Talier"
    ws_stat["A1"].font = Font(name="Arial", size=13, bold=True, color=HEADER_FONT)
    ws_stat["A1"].fill = PatternFill("solid", fgColor=HEADER_COLOR)
    ws_stat["A1"].alignment = Alignment(horizontal="center")
    ws_stat.merge_cells("A1:C1")

    stats_labels = [
        ("Total réservations", f"=COUNTA(Réservations!A3:A10000)"),
        ("Total couverts",     f"=SUMIF(Réservations!A3:A10000,\"<>\",Réservations!F3:F10000)"),
        ("Chiffre d'affaires total (FCFA)", f"=SUM(Réservations!I3:I10000)"),
        ("Sur place",          f"=COUNTIF(Réservations!G3:G10000,\"Sur place\")"),
        ("À emporter",         f"=COUNTIF(Réservations!G3:G10000,\"À emporter\")"),
        ("Traiteur",           f"=COUNTIF(Réservations!G3:G10000,\"Service traiteur\")"),
    ]
    for i, (label, formula) in enumerate(stats_labels, 3):
        ws_stat[f"A{i}"] = label
        ws_stat[f"B{i}"] = formula
        ws_stat[f"A{i}"].font = Font(name="Arial", size=10, bold=True)
        ws_stat[f"B{i}"].font = Font(name="Arial", size=10)
        ws_stat[f"B{i}"].alignment = Alignment(horizontal="right")
    ws_stat.column_dimensions["A"].width = 38
    ws_stat.column_dimensions["B"].width = 22

    wb.save(EXCEL_FILE)
    return wb


def add_reservation(reservation: dict) -> int:
    """
    Ajoute une réservation dans le fichier Excel.

    Returns:
        Numéro de réservation attribué.
    """
    if not os.path.exists(EXCEL_FILE):
        _create_workbook()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Réservations"]

    # Trouver la prochaine ligne libre et le N° de résa
    next_row = ws.max_row + 1
    resa_num = next_row - 2  # -2 pour titre + en-tête

    # Formater la commande en texte lisible
    commande_parts = []
    for item in reservation.get("commande", []):
        qte = item.get("quantite", 1)
        nom = item.get("nom", "")
        prix = item.get("prix", 0)
        if isinstance(prix, int):
            prix_str = f"{prix:,} FCFA".replace(",", " ")
        else:
            prix_str = f"{prix} FCFA"
        accomp = item.get("accompagnement", "")
        line = f"{qte}x {nom} ({prix_str})"
        if accomp:
            line += f" + {accomp}"
        commande_parts.append(line)
    commande_text = " | ".join(commande_parts) if commande_parts else "À préciser"

    type_service_labels = {
        "sur_place": "Sur place",
        "emporter": "À emporter",
        "traiteur": "Service traiteur"
    }
    service_label = type_service_labels.get(
        reservation.get("type_service", "sur_place"), "Sur place"
    )

    total = reservation.get("total", 0)
    row_data = [
        resa_num,
        reservation.get("date", ""),
        reservation.get("heure", ""),
        reservation.get("nom_client", ""),
        reservation.get("telephone", ""),
        reservation.get("nb_personnes", 1),
        service_label,
        commande_text,
        total,
        reservation.get("calendar_link", ""),
        datetime.now().strftime("%d/%m/%Y %H:%M"),
    ]

    ws.append(row_data)

    # Styler la nouvelle ligne
    is_even = (next_row % 2 == 0)
    bg_color = ROW_ALT_COLOR if is_even else ROW_BASE_COLOR
    for col_idx, cell in enumerate(ws[next_row], 1):
        cell.font = Font(name="Arial", size=9)
        cell.fill = PatternFill("solid", fgColor=bg_color)
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 8))
        if col_idx == 9:  # Colonne Total
            cell.number_format = '#,##0'
            cell.font = Font(name="Arial", size=9, bold=True)

    ws.row_dimensions[next_row].height = 18

    wb.save(EXCEL_FILE)
    return resa_num
